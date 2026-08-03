#!/usr/bin/env python3
"""Convert CSV or TSV data into a clean XLSX workbook."""

from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path

import chardet
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

_INTEGER_RE = re.compile(r"^[+-]?(?:0|[1-9]\d*)$")
_FLOAT_RE = re.compile(r"^[+-]?(?:(?:\d+\.\d*)|(?:\d*\.\d+)|(?:\d+))(?:[eE][+-]?\d+)?$")


def _decode_source(path: Path) -> str:
    """Decode a delimited text file without silently replacing bytes.

    Args:
        path: Source text file.

    Returns:
        Decoded text.

    Raises:
        UnicodeError: If the detected encoding cannot decode the file.
    """
    content = path.read_bytes()
    try:
        return content.decode("utf-8-sig")
    except UnicodeDecodeError:
        detected = chardet.detect(content)
        encoding = str(detected.get("encoding") or "").strip()
        if not encoding:
            raise UnicodeError("Unable to detect the source encoding.")
        return content.decode(encoding)


def _parse_scalar(value: str) -> str | int | float | bool | date | datetime | None:
    """Infer a conservative spreadsheet scalar from text.

    Args:
        value: Raw field text.

    Returns:
        A typed scalar suitable for openpyxl.
    """
    stripped = value.strip()
    if not stripped:
        return None
    if stripped.upper() in {"TRUE", "FALSE"}:
        return stripped.upper() == "TRUE"
    if _INTEGER_RE.fullmatch(stripped):
        unsigned = stripped.lstrip("+-")
        if len(unsigned) == 1 or not unsigned.startswith("0"):
            return int(stripped)
    if _FLOAT_RE.fullmatch(stripped) and any(char in stripped for char in ".eE"):
        return float(stripped)
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", stripped):
        try:
            return date.fromisoformat(stripped)
        except ValueError:
            pass
    if "T" in stripped:
        try:
            return datetime.fromisoformat(stripped.replace("Z", "+00:00"))
        except ValueError:
            pass
    return value


def _display_width(value: object) -> int:
    """Estimate a useful Excel column width for a value.

    Args:
        value: Cell value.

    Returns:
        Approximate display width in monospace characters.
    """
    text = "" if value is None else str(value)
    return sum(
        2 if unicodedata.east_asian_width(character) in {"W", "F"} else 1
        for character in text
    )


def convert_csv_to_xlsx(
    input_path: Path,
    output_path: Path,
    *,
    delimiter: str | None,
    sheet_name: str,
    has_header: bool,
) -> None:
    """Convert delimited text into a formatted workbook.

    Args:
        input_path: Source CSV or TSV file.
        output_path: Destination XLSX file.
        delimiter: Explicit delimiter or None for detection.
        sheet_name: Destination worksheet name.
        has_header: Whether the first source row is a header.

    Raises:
        ValueError: If the source is empty or has no columns.
    """
    text = _decode_source(input_path)
    if not text.strip():
        raise ValueError("Delimited input is empty.")
    if delimiter is None:
        try:
            dialect = csv.Sniffer().sniff(text[:8192], delimiters=",\t;|")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = "\t" if input_path.suffix.lower() == ".tsv" else ","

    rows = list(csv.reader(text.splitlines(), delimiter=delimiter))
    if not rows or not any(rows):
        raise ValueError("Delimited input contains no cells.")
    column_count = max(len(row) for row in rows)
    if column_count == 0:
        raise ValueError("Delimited input contains no columns.")
    normalized = [row + [""] * (column_count - len(row)) for row in rows]
    if has_header:
        headers = [
            cell.strip() or f"Column {index + 1}"
            for index, cell in enumerate(normalized[0])
        ]
        data_rows = normalized[1:]
    else:
        headers = [f"Column {index + 1}" for index in range(column_count)]
        data_rows = normalized

    seen: dict[str, int] = {}
    for index, header in enumerate(headers):
        seen[header] = seen.get(header, 0) + 1
        if seen[header] > 1:
            headers[index] = f"{header} {seen[header]}"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31]
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="334155")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10, color="1F2937")
    border = Border(bottom=Side(style="thin", color="CBD5E1"))

    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 24

    for row_index, row in enumerate(data_rows, start=2):
        for column_index, raw_value in enumerate(row, start=1):
            value = _parse_scalar(raw_value)
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(vertical="top")
            if isinstance(value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm"
            elif isinstance(value, date):
                cell.number_format = "yyyy-mm-dd"
            elif isinstance(value, float):
                cell.number_format = "#,##0.00"
            elif isinstance(value, int):
                cell.number_format = "#,##0"
            elif isinstance(value, str) and value.startswith("="):
                cell.data_type = "s"

    if data_rows:
        table = Table(
            displayName="ImportedData",
            ref=f"A1:{worksheet.cell(row=len(data_rows) + 1, column=column_count).coordinate}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    for column_index in range(1, column_count + 1):
        width = max(
            _display_width(worksheet.cell(row=row_index, column=column_index).value)
            for row_index in range(1, worksheet.max_row + 1)
        )
        worksheet.column_dimensions[
            worksheet.cell(row=1, column=column_index).column_letter
        ].width = min(max(width + 2, 10), 48)

    worksheet.auto_filter.ref = worksheet.dimensions
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    load_workbook(output_path, read_only=True).close()


def main() -> int:
    """Run the command-line converter.

    Returns:
        Process exit code.
    """
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path, help="Source CSV or TSV file")
    parser.add_argument("output", type=Path, help="Destination XLSX file")
    parser.add_argument(
        "--delimiter",
        choices=("comma", "tab", "semicolon", "pipe"),
        help="Explicit delimiter; the default is automatic detection",
    )
    parser.add_argument("--sheet-name", default="Data", help="Worksheet name")
    parser.add_argument(
        "--no-header",
        action="store_true",
        help="Treat the first source row as data",
    )
    args = parser.parse_args()

    input_path = args.input.expanduser().resolve(strict=False)
    output_path = args.output.expanduser().resolve(strict=False)
    if not input_path.is_file():
        parser.error(f"Input file not found: {input_path}")
    if output_path.suffix.lower() != ".xlsx":
        parser.error("Output path must end with .xlsx")
    if input_path == output_path:
        parser.error("Input and output paths must be different.")
    delimiter_map = {"comma": ",", "tab": "\t", "semicolon": ";", "pipe": "|"}
    try:
        convert_csv_to_xlsx(
            input_path,
            output_path,
            delimiter=delimiter_map.get(args.delimiter),
            sheet_name=args.sheet_name,
            has_header=not args.no_header,
        )
    except Exception as exc:
        parser.error(str(exc))
    print(f"created: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
