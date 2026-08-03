#!/usr/bin/env python3
"""Inspect workbook or delimited-text structure and print compact JSON."""

from __future__ import annotations

import argparse
import csv
import json
import zipfile
from pathlib import Path

import chardet
import xlrd
from openpyxl import load_workbook

_EXCEL_ERRORS = {
    "#NULL!",
    "#DIV/0!",
    "#VALUE!",
    "#REF!",
    "#NAME?",
    "#NUM!",
    "#N/A",
    "#GETTING_DATA",
}


def _inspect_delimited(path: Path, sample_rows: int, sample_cols: int) -> dict:
    """Inspect a CSV or TSV file.

    Args:
        path: Source file.
        sample_rows: Maximum sample rows.
        sample_cols: Maximum sample columns.

    Returns:
        JSON-serializable inspection data.
    """
    content = path.read_bytes()
    detected = chardet.detect(content)
    encoding = str(detected.get("encoding") or "utf-8")
    text = content.decode(encoding)
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",\t;|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    rows = list(csv.reader(text.splitlines(), delimiter=delimiter))
    return {
        "kind": "delimited",
        "encoding": encoding,
        "delimiter": delimiter,
        "rows": len(rows),
        "columns": max((len(row) for row in rows), default=0),
        "sample": [row[:sample_cols] for row in rows[:sample_rows]],
    }


def _inspect_xls(path: Path, sample_rows: int, sample_cols: int) -> dict:
    """Inspect a legacy XLS workbook.

    Args:
        path: XLS file.
        sample_rows: Maximum sample rows per sheet.
        sample_cols: Maximum sample columns per sheet.

    Returns:
        JSON-serializable inspection data.
    """
    workbook = xlrd.open_workbook(path, on_demand=True)
    sheets = []
    for name in workbook.sheet_names():
        sheet = workbook.sheet_by_name(name)
        sheets.append(
            {
                "name": name,
                "rows": sheet.nrows,
                "columns": sheet.ncols,
                "sample": [
                    sheet.row_values(row, end_colx=min(sheet.ncols, sample_cols))
                    for row in range(min(sheet.nrows, sample_rows))
                ],
            }
        )
    workbook.release_resources()
    return {"kind": "xls", "sheets": sheets, "formula_visibility": "unavailable"}


def _inspect_xlsx(path: Path, sample_rows: int, sample_cols: int) -> dict:
    """Inspect an Open XML workbook.

    Args:
        path: XLSX-family file.
        sample_rows: Maximum sample rows per sheet.
        sample_cols: Maximum sample columns per sheet.

    Returns:
        JSON-serializable inspection data.
    """
    keep_vba = path.suffix.lower() in {".xlsm", ".xltm"}
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=False,
        keep_vba=keep_vba,
    )
    sheets = []
    for sheet in workbook.worksheets:
        formulas = 0
        errors = 0
        nonempty = 0
        sample = []
        for row_index, row in enumerate(sheet.iter_rows()):
            values = []
            for cell in row:
                value = cell.value
                if value is not None:
                    nonempty += 1
                if cell.data_type == "f":
                    formulas += 1
                if cell.data_type == "e" or value in _EXCEL_ERRORS:
                    errors += 1
                if row_index < sample_rows and len(values) < sample_cols:
                    values.append(value)
            if row_index < sample_rows:
                sample.append(values)
        sheets.append(
            {
                "name": sheet.title,
                "state": sheet.sheet_state,
                "rows": sheet.max_row,
                "columns": sheet.max_column,
                "nonempty_cells": nonempty,
                "formula_cells": formulas,
                "error_cells": errors,
                "sample": sample,
            }
        )
    properties = {
        "title": workbook.properties.title,
        "creator": workbook.properties.creator,
        "last_modified_by": workbook.properties.lastModifiedBy,
    }
    workbook.close()
    with zipfile.ZipFile(path) as package:
        names = package.namelist()
    return {
        "kind": "openxml-workbook",
        "properties": properties,
        "sheets": sheets,
        "tables": sum(name.startswith("xl/tables/table") for name in names),
        "charts": sum(name.startswith("xl/charts/chart") for name in names),
        "drawings": sum(name.startswith("xl/drawings/drawing") for name in names),
        "external_links": sum(name.startswith("xl/externalLinks/") for name in names),
        "contains_vba": "xl/vbaProject.bin" in names,
    }


def main() -> int:
    """Run the workbook inspector.

    Returns:
        Process exit code.
    """
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path, help="Workbook, CSV, or TSV file")
    parser.add_argument("--sample-rows", type=int, default=6)
    parser.add_argument("--sample-cols", type=int, default=10)
    args = parser.parse_args()
    path = args.input.expanduser().resolve(strict=False)
    if not path.is_file():
        parser.error(f"Input file not found: {path}")
    if args.sample_rows < 0 or args.sample_cols < 0:
        parser.error("Sample limits must be non-negative.")

    try:
        suffix = path.suffix.lower()
        if suffix in {".csv", ".tsv"}:
            result = _inspect_delimited(path, args.sample_rows, args.sample_cols)
        elif suffix == ".xls":
            result = _inspect_xls(path, args.sample_rows, args.sample_cols)
        elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            result = _inspect_xlsx(path, args.sample_rows, args.sample_cols)
        else:
            raise ValueError(f"Unsupported spreadsheet format: {suffix or '(none)'}")
    except Exception as exc:
        parser.error(str(exc))
    result["path"] = str(path)
    result["size_bytes"] = path.stat().st_size
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
