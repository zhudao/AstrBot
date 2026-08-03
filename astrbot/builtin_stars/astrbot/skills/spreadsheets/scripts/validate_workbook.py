#!/usr/bin/env python3
"""Validate spreadsheet structure and common formula failures."""

from __future__ import annotations

import argparse
import csv
import json
import zipfile
from pathlib import Path

import xlrd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

_EXCEL_ERRORS = {
    "#NULL!",
    "#DIV/0!",
    "#VALUE!",
    "#REF!",
    "#NAME?",
    "#NUM!",
    "#N/A",
    "#GETTING_DATA",
}


def _validate_delimited(path: Path, errors: list[str], warnings: list[str]) -> dict:
    """Validate row widths in a CSV or TSV file.

    Args:
        path: Source file.
        errors: Error accumulator.
        warnings: Warning accumulator.

    Returns:
        Validation statistics.
    """
    text = path.read_text(encoding="utf-8-sig")
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",\t;|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    rows = list(csv.reader(text.splitlines(), delimiter=delimiter))
    if not rows:
        errors.append("The file contains no rows.")
        return {"rows": 0, "columns": 0}
    widths = [len(row) for row in rows]
    expected = max(set(widths), key=widths.count)
    inconsistent = [
        index + 1 for index, width in enumerate(widths) if width != expected
    ]
    if inconsistent:
        warnings.append(
            "Rows with a non-modal column count: "
            + ", ".join(map(str, inconsistent[:20]))
        )
    return {"rows": len(rows), "columns": max(widths, default=0)}


def _validate_xlsx(path: Path, errors: list[str], warnings: list[str]) -> dict:
    """Validate an XLSX-family workbook.

    Args:
        path: Workbook path.
        errors: Error accumulator.
        warnings: Warning accumulator.

    Returns:
        Validation statistics.
    """
    with zipfile.ZipFile(path) as package:
        bad_member = package.testzip()
        names = package.namelist()
    if bad_member:
        errors.append(f"Corrupt ZIP member: {bad_member}")

    keep_vba = path.suffix.lower() in {".xlsm", ".xltm"}
    workbook = load_workbook(
        path,
        read_only=False,
        data_only=False,
        keep_vba=keep_vba,
    )
    if not workbook.worksheets:
        errors.append("Workbook contains no worksheets.")
    if workbook.worksheets and all(
        sheet.sheet_state != "visible" for sheet in workbook.worksheets
    ):
        errors.append("Workbook has no visible worksheet.")

    formula_count = 0
    error_count = 0
    table_names: set[str] = set()
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if cell.data_type == "f":
                    formula_count += 1
                    if "#REF!" in str(value).upper():
                        errors.append(
                            f"Broken formula reference at '{sheet.title}'!{cell.coordinate}"
                        )
                if cell.data_type == "e" or value in _EXCEL_ERRORS:
                    error_count += 1
                    errors.append(
                        f"Excel error {value!r} at '{sheet.title}'!{cell.coordinate}"
                    )

        for table in sheet.tables.values():
            if table.name in table_names:
                errors.append(f"Duplicate table name: {table.name}")
            table_names.add(table.name)
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            if (
                min_row < 1
                or min_col < 1
                or max_row > sheet.max_row
                or max_col > sheet.max_column
            ):
                errors.append(f"Table {table.name} has an invalid range: {table.ref}")

        for chart in sheet._charts:
            for series in chart.ser:
                for reference in (
                    getattr(getattr(series, "val", None), "numRef", None),
                    getattr(getattr(series, "cat", None), "numRef", None),
                    getattr(getattr(series, "cat", None), "strRef", None),
                ):
                    formula = getattr(reference, "f", None)
                    if formula and "#REF!" in formula.upper():
                        errors.append(
                            f"Chart in '{sheet.title}' contains a broken source range."
                        )

    contains_vba = "xl/vbaProject.bin" in names
    if keep_vba and not contains_vba:
        warnings.append(
            "Macro-enabled extension is present but no VBA project was found."
        )
    if not keep_vba and contains_vba:
        errors.append("VBA content exists in a non-macro workbook extension.")
    external_links = sum(name.startswith("xl/externalLinks/") for name in names)
    if external_links:
        warnings.append(f"Workbook contains {external_links} external link part(s).")
    workbook.close()
    return {
        "sheets": len(workbook.worksheets),
        "formulas": formula_count,
        "error_cells": error_count,
        "tables": len(table_names),
        "external_links": external_links,
        "contains_vba": contains_vba,
    }


def main() -> int:
    """Run workbook validation.

    Returns:
        Zero for a valid workbook, otherwise one.
    """
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path, help="Workbook, CSV, or TSV file")
    args = parser.parse_args()
    path = args.input.expanduser().resolve(strict=False)
    if not path.is_file():
        parser.error(f"Input file not found: {path}")

    errors: list[str] = []
    warnings: list[str] = []
    try:
        suffix = path.suffix.lower()
        if suffix in {".csv", ".tsv"}:
            statistics = _validate_delimited(path, errors, warnings)
        elif suffix == ".xls":
            workbook = xlrd.open_workbook(path, on_demand=True)
            statistics = {"sheets": workbook.nsheets, "format": "xls-read-only"}
            workbook.release_resources()
        elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            statistics = _validate_xlsx(path, errors, warnings)
        else:
            raise ValueError(f"Unsupported spreadsheet format: {suffix or '(none)'}")
    except Exception as exc:
        errors.append(str(exc))
        statistics = {}

    result = {
        "valid": not errors,
        "path": str(path),
        "statistics": statistics,
        "errors": errors,
        "warnings": warnings,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    return 0 if not errors else 1


if __name__ == "__main__":
    raise SystemExit(main())
